import pandas as pd
from openpyxl import load_workbook

# Load the data
excel_file = 'Abyssal_Ship_Hardware.xlsx'
weapons = pd.read_excel(excel_file, sheet_name='Weapons')
ships = pd.read_excel(excel_file, sheet_name='Ships')
components = pd.read_excel(excel_file, sheet_name='Components')

# Function to handle updates based on type
def update_excel(data, df):
    new_data_df = pd.DataFrame([data])
    updated_df = pd.concat([df, new_data_df], ignore_index=True)
    return updated_df

# Example bulk updates
bulk_updates = [

{
    "data": {
        "Weapon Name": "SMS-1",
        "Type": "Short Range Missile",
        "Range Category": "Short Range",
        "Damage": "2d8",
        "Penetration": 8,
        "EWAR": 3,
        "Rate of Fire": "Single",
        "Special Properties": "Compact size, capable of engaging multiple targets with quick reloading.",
        "Component Size": 1,
        "Rating": "14",
        "Description": "The SMS-1 is a compact short-range missile system designed for quick and efficient engagements. It delivers 2d8 damage with moderate penetration, making it effective against lightly armored targets or for point defense. The system is designed for rapid reloading, allowing for quick follow-up shots in combat situations."
    },
    "type": "weapon"
},


{
    "data": {
        "Weapon Name": "LCGC-7",
        "Type": "Large Calibre Gauss Cannon",
        "Range Category": "Long Range",
        "Damage": "4d10",
        "Penetration": 15,
        "EWAR": 6,
        "Rate of Fire": "Sustained",
        "Special Properties": "High penetration, effective against heavily armored targets and capital ships.",
        "Component Size": 3,
        "Rating": "27",
        "Description": "The LCGC-7 is a powerful long-range weapon system designed to punch through heavy armor with ease. Capable of dealing massive damage with 4d10 and high penetration, it's ideal for engaging large vessels or fortified structures. Its sustained rate of fire ensures it can maintain pressure on targets over extended engagements."
    },
    "type": "weapon"
},


{
    "data": {
        "Weapon Name": "PDGT-4",
        "Type": "Point Defense Gauss Weapon",
        "Range Category": "Short Range",
        "Damage": "2d6",
        "Penetration": 6,
        "EWAR": 4,
        "Rate of Fire": "Rapid Fire",
        "Special Properties": "High rate of fire, optimized for intercepting incoming missiles and small craft.",
        "Component Size": 1,
        "Rating": "16",
        "Description": "The PDGT-4 is a rapid-fire Gauss weapon system designed for point defense. It excels at intercepting incoming missiles, projectiles, and small enemy craft, making it an essential defensive weapon for any ship. Its high rate of fire ensures that threats are neutralized quickly and efficiently."
    },
    "type": "weapon"
},


{
    "data": {
        "Weapon Name": "HIPL-9",
        "Type": "High Intensity Laser",
        "Range Category": "Medium Range",
        "Damage": "3d8",
        "Penetration": 12,
        "EWAR": 8,
        "Rate of Fire": "Burst",
        "Special Properties": "Capable of precise strikes, effective against both shields and armor.",
        "Component Size": 2,
        "Rating": "21",
        "Description": "The HIPL-9 is a versatile medium-range laser weapon, capable of delivering powerful burst attacks that can penetrate shields and armor alike. Its precision targeting and high damage output make it a formidable choice for both offensive and defensive engagements."
    },
    "type": "weapon"
},


{
    "data": {
        "Weapon Name": "TLS-12",
        "Type": "Torpedo",
        "Range Category": "Long Range",
        "Damage": "5d12",
        "Penetration": 18,
        "EWAR": 10,
        "Rate of Fire": "Single",
        "Special Properties": "Heavy damage, capable of delivering devastating blows to large targets.",
        "Component Size": 3,
        "Rating": "30",
        "Description": "The TLS-12 is a heavy-duty torpedo launcher system designed for long-range engagements. It delivers devastating damage with 5d12 and very high penetration, making it the perfect weapon for crippling or destroying capital ships and heavily fortified installations. Its single-shot capacity requires careful targeting, but the impact is well worth it."
    },
    "type": "weapon"
},

]

# Process each update
for update in bulk_updates:
    if update['type'] == 'weapon':
        weapons = update_excel(update['data'], weapons)
    elif update['type'] == 'ship':
        ships = update_excel(update['data'], ships)
    elif update['type'] == 'component':
        components = update_excel(update['data'], components)
    else:
        print(f"Unknown data type: {update['type']}")

# Load the existing workbook
workbook = load_workbook(excel_file)
# Select the sheets to update
weapons_sheet = workbook['Weapons']
ships_sheet = workbook['Ships']
components_sheet = workbook['Components']

# Update the Weapons sheet
for index, row in weapons.iterrows():
    for col_num, value in enumerate(row, 1):
        weapons_sheet.cell(row=index + 2, column=col_num, value=value)  # +2 to skip the header

# Similarly, update the Ships and Components sheets if needed
# Ships and Components updates follow the same pattern as the Weapons update

# Save the workbook to keep formatting
workbook.save(excel_file)

print("Bulk updates completed and saved to the Excel file.")
